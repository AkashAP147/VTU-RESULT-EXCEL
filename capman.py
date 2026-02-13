import os
from selenium import webdriver
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


# ==========================================
# SEMESTER URLS
# ==========================================
SEM_URLS = {
    "1": "https://results.vtu.ac.in/DJcbcs24/index.php",
    "2": "https://results.vtu.ac.in/JJEcbcs24/index.php",
    "3": "https://results.vtu.ac.in/DJcbcs25/index.php",
    "4": "https://results.vtu.ac.in/JJEcbcs25/index.php",
    "5": "https://results.vtu.ac.in/D25J26Ecbcs/index.php"
}


# ==========================================
# SUBJECT MAPS PER SEMESTER
# ==========================================
SEM_SUBJECT_MAPS = {
    "1": {
        "MATHEMATICS FOR CSE STREAM-I": "MATHS",
        "PHYSICS FOR CSE STREAM": "PHY",
        "PRINCIPLES OF PROGRAMMING USING C": "C",
        "COMMUNICATIVE ENGLISH": "ENG",
        "INDIAN CONSTITUTION": "IC",
        "INNOVATION AND DESIGN THINKING": "IDT",
        "INTRODUCTION TO CIVIL ENGINEERING": "CIVIL",
        "RENEWABLE ENERGY SOURCES": "RES",
    },
    "2": {
        "MATHEMATICS-II FOR CSE STREAM": "MATHS2",
        "APPLIED CHEMISTRY FOR CSE STREAM": "CHEM",
        "COMPUTER-AIDED ENGINEERING DRAWING": "CAED",
        "PROFESSIONAL WRITING SKILLS IN ENGLISH": "PWSE",
        "SAMSKRUTIKA KANNADA": "SK",
        "SCIENTIFIC FOUNDATIONS OF HEALTH": "SFH",
        "INTRODUCTION TO PYTHON PROGRAMMING": "PY",
        "INTRODUCTION TO ELECTRONICS COMMUNICATION": "ELC",
    },
    "3": {
        "MATHEMATICS FOR COMPUTER SCIENCE": "M3",
        "DIGITAL DESIGN & COMPUTER ORGANIZATION": "DDCO",
        "OPERATING SYSTEMS": "OS",
        "DATA STRUCTURES AND APPLICATIONS": "DSA",
        "DATA STRUCTURES LAB": "DSL",
        "SOCIAL CONNECT AND RESPONSIBILITY": "SCR",
        "NATIONAL SERVICE SCHEME": "NSS",
        "DATA ANALYTICS WITH EXCEL": "DAE",
        "OBJECT ORIENTED PROGRAMMING WITH JAVA": "OOPJ",
    },
    "4": {
        "ANALYSIS & DESIGN OF ALGORITHMS": "ADA",
        "ARTIFICIAL INTELLIGENCE": "AI",
        "DATABASE MANAGEMENT SYSTEMS": "DBMS",
        "ANALYSIS & DESIGN OF ALGORITHMS LAB": "ADAL",
        "BIOLOGY FOR COMPUTER ENGINEERS": "BIO",
        "UNIVERSAL HUMAN VALUES COURSE": "UHV",
        "NATIONAL SERVICE SCHEME": "NSS",
        "DISCRETE MATHEMATICAL STRUCTURES": "DMS",
        "TECHNICAL WRITING USING LATEX LAB": "TWL",
    },
    "5": {
        "SOFTWARE ENGINEERING AND PROJECT MANAGEMENT": "SEPM",
        "COMPUTER NETWORKS": "CN",
        "THEORY OF COMPUTATION": "TOC",
        "DATA VISUALIZATION LAB": "DVL",
        "MINI PROJECT": "MINI",
        "RESEARCH METHODOLOGY AND IPR": "RMIPR",
        "ENVIRONMENTAL STUDIES AND E-WASTE MANAGEMENT": "EVS",
        "NATIONAL SERVICE SCHEME": "NSS",
        "UNIX SYSTEM PROGRAMMING": "UNIX",
    },
}


# ==========================================
# EXCEL SETUP
# ==========================================
def setup_excel(file_name, sem):
    subject_short = list(SEM_SUBJECT_MAPS[sem].values())
    headers = ["USN", "Student Name"] + subject_short + ["TOTAL", "PERCENTAGE"]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        # Check if existing headers match the selected semester
        existing_headers = [ws.cell(row=1, column=c).value for c in range(1, len(headers) + 1)]
        if existing_headers != headers:
            print(f"⚠ Existing file has different headers. Overwriting row 1 with Sem {sem} headers.")
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            wb.save(file_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_name)
    return wb, ws


# ==========================================
# SAVE DATA
# ==========================================
def save_student(ws, wb, file_name, usn, name, marks):
    total = sum(int(m) for m in marks.values() if m != "")
    count = sum(1 for m in marks.values() if m != "")
    percentage = round((total / (count * 100)) * 100, 2) if count > 0 else 0

    ws.append([usn, name] + list(marks.values()) + [total, percentage])
    wb.save(file_name)


# ==========================================
# EXTRACT RESULT
# ==========================================
def extract_result(html, sem):
    soup = BeautifulSoup(html, "html.parser")

    if "University Seat Number" not in soup.text:
        return None, None, None

    usn = ""
    name = ""

    tables = soup.find_all("table")
    for table in tables:
        for row in table.find_all("tr"):
            cols = row.find_all("td")
            if len(cols) >= 2:
                label = cols[0].text.strip()
                value = cols[1].text.strip()
                if "University Seat Number" in label:
                    usn = value.replace(":", "").strip()
                if "Student Name" in label:
                    name = value.replace(":", "").strip()

    subject_map = SEM_SUBJECT_MAPS[sem]
    subject_short = list(subject_map.values())
    marks = {s: "" for s in subject_short}

    div_rows = soup.find_all("div", class_="divTableRow")
    for row in div_rows:
        cols = row.find_all("div", class_="divTableCell")
        if len(cols) == 7 and cols[0].text.strip() != "Subject Code":
            subject_name = cols[1].text.strip().upper()
            total_marks = cols[4].text.strip()

            for full, short in subject_map.items():
                if full in subject_name:
                    marks[short] = total_marks

    return usn, name, marks


# ==========================================
# NEXT USN
# ==========================================
def next_usn(usn):
    prefix = usn[:-3]
    number = int(usn[-3:])
    return prefix + str(number + 1).zfill(3)


# ==========================================
# MAIN
# ==========================================
def main():

    print("\n📚 VTU Result Scraper - Select Semester")
    print("=" * 50)
    print("1. Semester 1 (DJcbcs24)")
    print("2. Semester 2 (JJEcbcs24)")
    print("3. Semester 3 (DJcbcs25)")
    print("4. Semester 4 (JJEcbcs25)")
    print("5. Semester 5 (D25J26Ecbcs)")
    print("=" * 50)

    sem = input("Enter Semester (1-5): ").strip()

    if sem not in SEM_URLS:
        print("❌ Invalid semester!")
        return

    VTU_URL = SEM_URLS[sem]

    default_name = f"VTU_Sem{sem}_Results"
    excel_name = input(f"Enter Excel file name (default: {default_name}): ").strip()
    if not excel_name:
        excel_name = default_name
    EXCEL_FILE = excel_name + ".xlsx"

    start_usn = input("Enter Starting USN: ").strip()

    wb, ws = setup_excel(EXCEL_FILE, sem)

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    wait = WebDriverWait(driver, 600)

    current_usn = start_usn

    print(f"\n🚀 VTU Result Scraper Started")
    print(f"📚 Semester: {sem} → {VTU_URL}")
    print(f"📊 Excel: {EXCEL_FILE}")
    print(f"🎯 Starting USN: {start_usn}")
    print("Solve CAPTCHA manually → Submit → Auto next USN")
    print("Press Ctrl + C to stop\n")

    try:
        while True:

            driver.get(VTU_URL)

            wait.until(EC.presence_of_element_located((By.NAME, "lns")))

            usn_box = driver.find_element(By.NAME, "lns")
            usn_box.clear()
            usn_box.send_keys(current_usn)

            print(f"➡ Checking: {current_usn}")
            print("Solve CAPTCHA and click Submit...")

            try:
                wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//*[contains(text(),'University Seat Number')]")
                    )
                )
            except TimeoutException:
                print("❌ Timeout — moving to next USN")
                current_usn = next_usn(current_usn)
                continue
            except UnexpectedAlertPresentException:
                alert = driver.switch_to.alert
                print(f"⚠ {alert.text}")
                alert.accept()
                current_usn = next_usn(current_usn)
                continue

            html = driver.page_source
            usn, name, marks = extract_result(html, sem)

            if usn:
                print(f"📋 {usn} - {name}")
                for sub, val in marks.items():
                    print(f"   {sub}: {val}")
                save_student(ws, wb, EXCEL_FILE, usn, name, marks)
                print(f"✔ Saved to Excel")
            else:
                print(f"❌ Could not extract result for {current_usn}")

            current_usn = next_usn(current_usn)

    except KeyboardInterrupt:
        print("\n🛑 Stopped")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()