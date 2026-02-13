import os
import re
from selenium import webdriver
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


# ==========================================
# SEMESTER URLS
# ==========================================

SEM_URLS = {
    "1": "https://results.vtu.ac.in/DJcbcs24/resultpage.php",
    "2": "https://results.vtu.ac.in/JJEcbcs24/resultpage.php",
    "3": "https://results.vtu.ac.in/DJcbcs25/resultpage.php",
    "4": "https://results.vtu.ac.in/JJEcbcs25/resultpage.php",
    "5": "https://results.vtu.ac.in/D25J26Ecbcs/index.php"
}


# ==========================================
# CREATE SHORT NAME
# ==========================================

def create_short_name(subject):

    ignore = {"AND", "OF", "THE", "FOR", "WITH", "&"}

    words = re.split(r'\s+', subject.upper())
    short = ""

    for word in words:
        if word not in ignore and word.isalpha():
            short += word[0]

    return short[:5]  # limit length


# ==========================================
# NEXT USN
# ==========================================

def next_usn(usn):
    prefix = usn[:-3]
    number = int(usn[-3:])
    return prefix + str(number + 1).zfill(3)


# ==========================================
# EXTRACT RESULT (AUTO SUBJECT DETECT)
# ==========================================

def extract_result(html):

    soup = BeautifulSoup(html, "html.parser")

    if "University Seat Number" not in soup.text:
        return None, None, None, None

    usn = ""
    name = ""

    tables = soup.find_all("table")
    for table in tables:
        for row in table.find_all("tr"):
            cols = row.find_all("td")
            if len(cols) >= 2:
                if "University Seat Number" in cols[0].text:
                    usn = cols[1].text.strip()
                if "Student Name" in cols[0].text:
                    name = cols[1].text.strip()

    subjects = {}

    rows = soup.find_all("div", class_="divTableRow")
    for row in rows:
        cols = row.find_all("div", class_="divTableCell")

        if len(cols) == 7 and cols[0].text.strip() != "Subject Code":

            subject_name = cols[1].text.strip().upper()
            total_marks = cols[4].text.strip()

            short = create_short_name(subject_name)

            subjects[short] = total_marks

    # Sort subjects alphabetically
    sorted_subjects = dict(sorted(subjects.items()))

    return usn, name, sorted_subjects, list(sorted_subjects.keys())


# ==========================================
# MAIN
# ==========================================

def main():

    sem = input("Select Semester (1-5): ").strip()
    if sem not in SEM_URLS:
        print("Invalid semester")
        return

    VTU_URL = SEM_URLS[sem]

    excel_name = input("Enter Excel file name: ").strip() + ".xlsx"
    start_usn = input("Enter Starting USN: ").strip()

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    wait = WebDriverWait(driver, 600)

    wb = None
    ws = None
    headers_created = False

    current_usn = start_usn

    try:
        while True:

            driver.get(VTU_URL)

            wait.until(EC.presence_of_element_located((By.NAME, "lns")))
            driver.find_element(By.NAME, "lns").clear()
            driver.find_element(By.NAME, "lns").send_keys(current_usn)

            print(f"Checking: {current_usn}")
            print("Solve CAPTCHA and click Submit...")

            try:
                wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//*[contains(text(),'University Seat Number')]")
                    )
                )
            except TimeoutException:
                current_usn = next_usn(current_usn)
                continue
            except UnexpectedAlertPresentException:
                driver.switch_to.alert.accept()
                current_usn = next_usn(current_usn)
                continue

            html = driver.page_source
            usn, name, subjects, subject_headers = extract_result(html)

            if usn:

                if not headers_created:
                    HEADERS = ["USN", "Student Name"] + subject_headers + ["TOTAL", "PERCENTAGE"]

                    if os.path.exists(excel_name):
                        wb = load_workbook(excel_name)
                        ws = wb.active
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.append(HEADERS)

                    headers_created = True

                total = sum(int(v) for v in subjects.values() if v.isdigit())
                count = len(subjects)
                percentage = round((total / (count * 100)) * 100, 2)

                ws.append([usn, name] + list(subjects.values()) + [total, percentage])
                wb.save(excel_name)

                print(f"Saved: {usn}")

            current_usn = next_usn(current_usn)

    except KeyboardInterrupt:
        print("Stopped by user")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
